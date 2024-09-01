import csv
import json
import os
import re
import time
import requests
import sys
import logging
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tqdm import tqdm
from ratelimiter import RateLimiter
import undetected_chromedriver as uc
import urllib.parse
from selenium.webdriver.common.by import By

LEETCODE_QUERY = '''
https://leetcode.com/graphql?query=query
{     
      userContestRanking(username:  "{<username>}") 
      {
        attendedContestsCount
        rating
        globalRanking
        totalParticipants
        topPercentage    
      }
}
'''

MAX_REQUESTS_PER_SECOND = 2


class Participant:
    handle = ""
    geeksforgeeks_handle = ""
    codeforces_handle = ""
    leetcode_handle = ""
    codechef_handle = ""
    hackerrank_handle = ""
    geeksforgeeks_url_exists = False
    codeforces_url_exists = False
    leetcode_url_exists = False
    codechef_url_exists = False
    hackerrank_url_exists = False
    def __init__(self, handle, geeksforgeeks_handle, codeforces_handle, leetcode_handle, codechef_handle,
                 hackerrank_handle, geeksforgeeks_url_exists=False, codeforces_url_exists=False, leetcode_url_exists=False, 
                    codechef_url_exists=False, hackerrank_url_exists=False):
        handle = remove_non_ascii(handle)
        geeksforgeeks_handle = remove_non_ascii(geeksforgeeks_handle)
        codeforces_handle = remove_non_ascii(codeforces_handle)
        leetcode_handle = remove_non_ascii(leetcode_handle)
        codechef_handle = remove_non_ascii(codechef_handle)
        hackerrank_handle = remove_non_ascii(hackerrank_handle)
        # remove @ from the leeetcode handle
        hackerrank_handle = hackerrank_handle.replace('@', '')
        leetcode_handle = leetcode_handle.replace('@', '')
        geeksforgeeks_handle = geeksforgeeks_handle.strip()
        self.handle = handle
        self.geeksforgeeks_handle = geeksforgeeks_handle
        self.codeforces_handle = codeforces_handle
        self.leetcode_handle = leetcode_handle
        self.codechef_handle = codechef_handle
        self.hackerrank_handle = hackerrank_handle
        self.geeksforgeeks_url_exists = geeksforgeeks_url_exists
        self.codeforces_url_exists = codeforces_url_exists
        self.leetcode_url_exists = leetcode_url_exists
        self.codechef_url_exists = codechef_url_exists
        self.hackerrank_url_exists = hackerrank_url_exists


def remove_non_ascii(input_string):
    return re.sub(r'[\t\n\x0B\f\r]+', '', input_string)


def check_url_exists(url):
    # if url is leeetcode
    if "https://leetcode.com/" in url:
        try:
            response = requests.get(url)
            if response.status_code == 200:
                # read response as json
                response_json = response.json()
                # if the response contains the key "errors", then the handle does not exist
                if response_json.get("errors"):
                    return False, response.url
                return True, response.url
            return False, response.url
        except requests.exceptions.RequestException:
            return False, "Exception"
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "TE": "Trailers"
    }
    # if url is hackerrank
    if "https://www.hackerrank.com/" in url:
        try:
            response = requests.get(url, headers=header)
            soup = BeautifulSoup(response.text, 'html.parser')
            # Extract the title of the page
            title = soup.title.string
            print(title)
            # Hackerrank handles that do not exist have the title "HTTP 404: Page Not Found | HackerRank"
            # But the title in beautiful soup is "Programming Problems and Competitions :: HackerRank"
            # If user exists, title will be " Name - User Profile | HackerRank"
            
            return True, response.url
        except requests.exceptions.RequestException:
            return False, "Exception"
    try:
        response = requests.get(url, headers=header)
        if response.status_code == 200:
            # Check if the final URL is the same as the original URL (no redirect), if redirected, then URL does not
            # exist codeforces redirect is found by checking if final url is https://codeforces.com/ geeksforgeeks
            # redirect is found by checking if final url is
            # https://auth.geeksforgeeks.org/?to=https://auth.geeksforgeeks.org/profile.php codechef redirect is
            # found by checking if final url is https://www.codechef.com/ Hackerrank and Leetcode return 404 error if
            # handle does not exist
            logging.debug(f"Checking : {response.url}")
            if (response.url == "https://codeforces.com/" or response.url == ("https://auth.geeksforgeeks.org/?to=https"
                                                                              "://auth.geeksforgeeks.org/profile.php")
                    or response.url == "https://www.codechef.com/"):
                return False, response.url
            else:
                return True, response.url
        return False, response.url
    except requests.exceptions.RequestException:
        return False, "Exception"


def process_geeksforgeeks(participants):
    """
    Process GeeksForGeeks handles for each participant and log the progress.

    Args:
    participants (list): List of participant objects

    Returns:
    None
    """
    # Configure logging
    logging.basicConfig(filename='geeksforgeeks_debug.log', level=logging.DEBUG)

    # Initialize variables to store the last user's status and handle
    last_user_status = None
    last_user_handle = None

    # Iterate through each participant
    with tqdm(participants, desc="Processing GeeksForGeeks Handles", unit="participant") as pbar:
        for participant in pbar:
            # Check if GeeksForGeeks handle is valid
            if participant.geeksforgeeks_handle != '#N/A':
                logging.debug(f"Checking GeeksForGeeks URL for participant {participant.handle}")

                # Check if the GeeksForGeeks URL exists
                geeksforgeeks_url_exists, response_url = check_url_exists(
                    "https://auth.geeksforgeeks.org/user/" + participant.geeksforgeeks_handle)
                logging.debug(f"GeeksForGeeks URL exists: {geeksforgeeks_url_exists}, Response URL: {response_url}")

                # Retry if the GeeksForGeeks URL does not exist
                if not geeksforgeeks_url_exists and participant.geeksforgeeks_handle != '#N/A':
                    logging.debug(f"Retrying GeeksForGeeks URL check for participant {participant.handle}")
                    geeksforgeeks_url_exists, response_url = check_url_exists(
                        "https://auth.geeksforgeeks.org/user/" + participant.geeksforgeeks_handle)
                    logging.debug(f"GeeksForGeeks URL retry: {geeksforgeeks_url_exists}, Response URL: {response_url}")

                # Write participant data to file
                with open('geeksforgeeks_handles.txt', 'a') as file:
                    file.write(f"{participant.handle}, {participant.geeksforgeeks_handle}, {geeksforgeeks_url_exists}\n")
                logging.debug(
                    f"Data written to file for participant {participant.handle}: {participant.geeksforgeeks_handle},"
                    f" {geeksforgeeks_url_exists}")
                logging.debug("---------------------------------------------------")

                # Update the last user's status and handle
                last_user_status = geeksforgeeks_url_exists
                last_user_handle = participant.geeksforgeeks_handle

            # Display the last user's status alongside their username within the progress bar
            if last_user_handle is not None:
                pbar.set_postfix({"Last User": last_user_handle, "Status": last_user_status})

            pbar.update(1)

    # Shutdown logging
    logging.shutdown()


def process_leetcode(participants):
    """
    Process the LeetCode handles of participants.

    :param participants: A list of Participant objects containing their handles and LeetCode handles.
    :return: None

    This function processes the LeetCode handles of participants by making API requests to retrieve their contest ranking information. It uses a rate limiter to ensure a maximum of 2 requests per second. The function uses undetected-chromedriver to run in headless mode and performs the following steps:
    1. Configures logging.
    2. Creates chrome options and configures undetected-chromedriver.
    3. Logs in to GitHub using the provided username and password.
    4. Opens a new tab and navigates to the LeetCode login page.
    5. Authorizes the GitHub login if prompted.
    6. Iterates over the participants and retrieves their contest ranking information using the LeetCode API.
    7. Parses the JSON response and checks if the response contains any errors.
    8. Writes the participant's handle, LeetCode handle, and a boolean indicating if the response was successful to a file.

    Note: The function assumes that the LeetCode API query is defined in the LEETCODE_QUERY variable and the maximum number of requests per second is defined in the MAX_REQUESTS_PER_SECOND variable.

    Raises:
    - RuntimeError: If there is an error parsing the JSON response or getting the content for a participant.
    - Exception: If there is an error processing the LeetCode handle for a participant.
    """
    # Configure logging
    counter = 1
    size = len(participants)

    # Rate limit the function to a maximum of 2 requests per second
    limiter = RateLimiter(max_calls=MAX_REQUESTS_PER_SECOND, period=1)

    # Create chrome options
    options = uc.ChromeOptions()
    options.add_argument("--auto-open-devtools-for-tabs")

    # Configure undetected-chromedriver to run in headless mode
    driver = uc.Chrome(version_main=128, options=options)

    # Login to GitHub
    driver.get("https://github.com/login")
    time.sleep(5)
    login = driver.find_element(By.NAME, "login")
    password = driver.find_element(By.NAME, "password")
    signin_btn = driver.find_element(By.NAME, "commit")
    # load username from USERNAME env variable
    username = os.environ.get('USERNAME')
    # load password from PASSWORD env variable
    passwd = os.environ.get('PASSWD')
    login.send_keys(username)
    password.send_keys(passwd)
    signin_btn.click()

    # Open new tab to https://leetcode.com/accounts/login/
    driver.get('https://leetcode.com/accounts/github/login/?next=%2F')
    time.sleep(5)
    try:
        authorize_btn = driver.find_element(By.NAME, "authorize")
        authorize_btn.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error: {e}")

    for participant in participants:
        handle = participant.handle
        leetcode_handle = participant.leetcode_handle
        # Construct URL for API request
        encoded_leetcode_handle = urllib.parse.quote(leetcode_handle, safe='')
        url = LEETCODE_QUERY.replace("{<username>}", encoded_leetcode_handle)
        url = url.replace(" ", "%20")
        try:
            with limiter:
                driver.get(url)

                # Parse JSON response
                try:
                    json_content = driver.find_element(By.TAG_NAME, "pre").text
                    json_content = json.loads(json_content)
                except Exception as e:
                    raise RuntimeError(f"Error parsing JSON response for {handle} with LeetCode handle {leetcode_handle}: {e}")

                try:
                    # Check if the response contains error
                    if json_content.get("errors"):
                        with open('leetcode_handles.txt', 'a') as file:
                            file.write(f"{handle}, {leetcode_handle}, False\n")
                        print(f"( {counter} / {size} ) Data written to file for participant {handle}: {leetcode_handle}, False")
                        print("---------------------------------------------------")
                        counter += 1
                        continue
                    else:
                        with open('leetcode_handles.txt', 'a') as file:
                            file.write(f"{handle}, {leetcode_handle}, True\n")
                        print(f"( {counter} / {size} ) Data written to file for participant {handle}: {leetcode_handle}, True")
                        print("---------------------------------------------------")
                        counter += 1
                except (KeyError, TypeError) as e:
                    raise RuntimeError(f"Error getting content for {handle} with LeetCode handle {leetcode_handle}: {e}")
        except Exception as e:
            raise RuntimeError(f"Error processing LeetCode handle for {handle}: {e}")
            
# Load API_KEY and API_SECRET from environment variables
API_KEY = os.getenv('CODEFORCES_API_KEY')
API_SECRET = os.getenv('CODEFORCES_API_SECRET')
CODEFORCES_URL = 'https://codeforces.com/api/user.info'

# Function to check if Codeforces users exist
def generate_random_string(length=6):
    """Generate a random string of specified length."""
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

def generate_api_signature(rand, method_name, handles, time, secret):
    """Generate API signature."""
    parameters = f"apiKey={API_KEY}&handles={handles}&time={time}"
    to_hash = f"{rand}/{method_name}?{parameters}#{secret}"
    
    hash_bytes = hashlib.sha512(to_hash.encode('utf-8')).digest()
    return ''.join(f"{byte:02x}" for byte in hash_bytes)

def fetch_codeforces_data(handles):
    """Fetch Codeforces user data using the API."""
    current_time = int(time.time())
    rand = generate_random_string()
    api_sig = generate_api_signature(rand, "user.info", ';'.join(handles), current_time, API_SECRET)
    
    # Construct the request URL
    url = f"{CODEFORCES_URL}?handles={';'.join(handles)}&apiKey={API_KEY}&time={current_time}&apiSig={rand}{api_sig}"
    
    retry_count = 0
    max_retries = 10
    while retry_count < max_retries:
        try:
            response = requests.get(url)
            response.raise_for_status()
            
            # Print and return JSON response
            json_response = response.json()
            # Log the response
            logging.debug(json_response)
            return json_response
        except requests.RequestException as e:
            retry_count += 1
            print(f"Error fetching Codeforces data. Retrying attempt {retry_count}: {e}")
    
    raise Exception("Failed to fetch Codeforces data after multiple retries.")

# Function to process Codeforces handles
def process_codeforces(participants):
    logging.basicConfig(filename='codeforces_debug.log', level=logging.DEBUG)

    # Load Codeforces handles from participants
    handles = {participant.codeforces_handle.replace(" ", "") for participant in participants if participant.codeforces_handle != '#N/A' and "@" not in participant.codeforces_handle}
    
    # Initialize variables
    remaining_handles = set(handles)
    all_valid_handles = set()
    all_batches_successful = True

    # Split handles into batches of 300 and process them
    batches = [list(remaining_handles)[i:i + 300] for i in range(0, len(remaining_handles), 300)]
    logging.debug(f"Total handles: {len(handles)}, Total batches: {len(batches)}")
    for index, batch in enumerate(batches, start=1):
        logging.info(f"The content of the batch {index} is {batch}")

    for index, batch in enumerate(batches, start=1):
        current_batch_message = f"""

        =======================================================
        PROCESSING BATCH {index} OF {len(batches)}
        =======================================================
        """
        print(current_batch_message)

        logging.debug(f"Processing batch {index} of {len(batches)}")

        while True:
            # Fetch user information
            response_json = check_codeforces_users(batch)
            if response_json["status"] == "OK":
                users = response_json["result"]
                
                # Collect valid handles
                valid_handles = {user["handle"] for user in users}
                all_valid_handles.update(valid_handles)
                
                # Find and remove non-existent handles
                handles_to_remove = set(batch) - valid_handles
                remaining_handles -= handles_to_remove
                
                # Log removed handles
                if handles_to_remove:
                    logging.debug(f"Handles not found: {handles_to_remove}")
                
                break

            elif response_json["status"] == "FAILED" and response_json["comment"].startswith("handles:"):
                # Find and remove non-existent handle
                # Format: "handles: User with handle <username> not found"
                handles_to_remove = {re.search(r"User with handle (.+) not found", response_json["comment"]).group(1)}
                remaining_handles -= handles_to_remove
                batch.remove(handles_to_remove.pop())
                logging.warning(f"Handles not found: {handles_to_remove}")

            else:
                logging.error(f"API Error: {response_json.get('comment', 'Unknown error')}")
                all_batches_successful = False
                break
            
            logging.debug(f"Remaining handles: {remaining_handles}")
    
    # Write valid handles to file
    with open('codeforces_handles.txt', 'a') as file:
        for participant in participants:
            participant.codeforces_handle = participant.codeforces_handle.replace(" ", "")
            if participant.codeforces_handle in all_valid_handles:
                file.write(f"{participant.handle}, {participant.codeforces_handle}, {True}\n")
            else:
                file.write(f"{participant.handle}, {participant.codeforces_handle}, {False}\n")
    
    # Logging and printing final status
    if all_batches_successful:
        success_message = """
        
        =======================================================
        ALL BATCHES PROCESSED SUCCESSFULLY!
        =======================================================
        """
        print(success_message)
        logging.info(success_message.strip())
    else:
        failure_message = """
        
        =======================================================
        SOME BATCHES FAILED TO PROCESS. CHECK LOG FOR DETAILS.
        =======================================================
        """
        print(failure_message)
        logging.error(failure_message.strip())
    
    logging.shutdown()

def process_codechef(participants):
    """
    Process the CodeChef handles for the given participants and log the progress.

    Args:
    participants (list): List of Participant objects.

    Returns:
    None
    """
    logging.basicConfig(filename='codechef_debug.log', level=logging.DEBUG)

    # Initialize variables to store the last user's status and handle
    last_user_status = None
    last_user_handle = None

    with tqdm(total=len(participants), desc="Processing CodeChef Handles", unit="participant") as pbar:
        for participant in participants:
            # Check CodeChef URL for each participant
            logging.debug(f"Checking CodeChef URL for participant {participant.handle}")

            if participant.codechef_handle != '#N/A':
                # Check if CodeChef URL exists
                codechef_url_exists, response_url = check_url_exists(
                    "https://www.codechef.com/users/" + participant.codechef_handle)
                logging.debug(f"CodeChef URL exists: {codechef_url_exists}, Response URL: {response_url}")

                if not codechef_url_exists and participant.codechef_handle != '#N/A':
                    # Retry checking CodeChef URL
                    logging.debug(f"Retrying CodeChef URL check for participant {participant.handle}")
                    codechef_url_exists, response_url = check_url_exists(
                        "https://www.codechef.com/users/" + participant.codechef_handle)
                    logging.debug(f"CodeChef URL retry: {codechef_url_exists}, Response URL: {response_url}")

                # Write participant data to file codechef_url_exists
                with open('codechef_handles.txt', 'a') as file:
                    file.write(f"{participant.handle}, {participant.codechef_handle}, {True}\n")
                logging.debug(f"Data written to file for participant {participant.handle}: {participant.codechef_handle},"
                            f" {codechef_url_exists}")
                logging.debug("---------------------------------------------------")

                # Update the last user's status and handle
                last_user_status = codechef_url_exists 
                last_user_handle = participant.codechef_handle

            # Display the last user's status alongside their username within the progress bar
                if last_user_handle is not None:
                    pbar.set_postfix({"Last User": last_user_handle, "Status": last_user_status})

                pbar.update(1)

    logging.shutdown()

def process_hackerrank(participants):
    """
    Process the HackerRank handles for the given participants and log the debugging information.

    Args:
    participants (list): List of Participant objects

    Returns:
    None
    """
    # Configure logging
    logging.basicConfig(filename='hackerrank_debug.log', level=logging.DEBUG)

    # Initialize variables to store the last user's status and handle
    last_user_status = None
    last_user_handle = None

    # Iterate through the participants
    with tqdm(participants, desc="Processing HackerRank Handles", unit="participant") as pbar:
        for participant in pbar:
            # Log the participant's HackerRank URL check
            logging.debug(f"Checking HackerRank URL for participant {participant.handle}")

            # Check if the HackerRank URL exists
            if participant.hackerrank_handle != '#N/A': 

                hackerrank_url_exists, response_url = check_url_exists(
                    "https://www.hackerrank.com/profile/" + participant.hackerrank_handle)
                logging.debug(f"HackerRank URL exists: {hackerrank_url_exists}, Response URL: {response_url}")

                # Retry checking the HackerRank URL
                if not hackerrank_url_exists and participant.hackerrank_handle != '#N/A':
                    logging.debug(f"Retrying HackerRank URL check for participant {participant.handle}")
                    hackerrank_url_exists, response_url = check_url_exists(
                        "https://www.hackerrank.com/profile/" + participant.hackerrank_handle)
                    logging.debug(f"HackerRank URL retry: {hackerrank_url_exists}, Response URL: {response_url}")

                # Write data to file
                with open('hackerrank_handles.txt', 'a') as file:
                    file.write(f"{participant.handle}, {participant.hackerrank_handle}, {True}\n")
                logging.debug(f"Data written to file for participant {participant.handle}: {participant.hackerrank_handle},"
                            f" {hackerrank_url_exists}")
                logging.debug("---------------------------------------------------")

                # Update the last user's status and handle
                last_user_status = hackerrank_url_exists
                last_user_handle = participant.hackerrank_handle

            # Display the last user's status alongside their username within the progress bar
            if last_user_handle is not None:
                pbar.set_postfix({"Last User": last_user_handle, "Status": last_user_status})

            pbar.update(1)

    # Shutdown logging
    logging.shutdown()


def load_excel_sheet(excel_sheet_path):
    participants = []
    workbook = load_workbook(excel_sheet_path)
    sheet = workbook.active
    total_rows = sheet.max_row - 2
    count = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(x == 'None' for x in row):
            break
        handle, geeksforgeeks_handle, codeforces_handle, leetcode_handle, codechef_handle, hackerrank_handle = row
        print(f"( {count} / {total_rows} ) Loading participant {handle}")
        participants.append(
            Participant(handle, geeksforgeeks_handle, codeforces_handle, leetcode_handle, codechef_handle,
                        hackerrank_handle))
        count += 1
    print("Finished loading participants")
    return participants


def load_csv_sheet(csv_sheet_path):
    """
    Load participant data from a CSV sheet and return a list of Participant objects.

    Args:
    csv_sheet_path (str): The file path to the CSV sheet

    Returns:
    list: A list of Participant objects
    """
    participants = []
    with open(csv_sheet_path, 'r') as temp_file:
        temp_reader = csv.reader(temp_file)  # Create a temporary reader to count total rows
        total_rows = sum(1 for _ in temp_reader) - 2  # Calculate total rows in the CSV
        # close the temporary file
        temp_file.close()
    with open(csv_sheet_path, 'r') as file:
        reader = csv.reader(file)
        print(f"Total rows in CSV: {total_rows}")  # Print the total rows in the CSV
        count = 1
        for row in reader:
            if row[0] == "Roll number":  # Skip the header row
                continue
            if all(x == 'None' or x == '' for x in row):  # Stop if all cells in the row are empty
                break
            handle, geeksforgeeks_handle, codeforces_handle, leetcode_handle, codechef_handle, hackerrank_handle = row
            print(f"( {count} / {total_rows} ) Loading participant {handle}")  # Print progress
            participant = Participant(handle, geeksforgeeks_handle, codeforces_handle, leetcode_handle, codechef_handle,
                                      hackerrank_handle)  # Create Participant object
            participants.append(participant)  # Add Participant object to list
            count += 1
    print("Finished loading participants")
    file.close()
    return participants


def combine_results(participants):
    """
    Combines handle details from multiple files and writes them to a CSV file called participant_details.csv.
    Each file contains handle details for a different platform, and the function loops through each file,
    reads the handle details, and writes them to the CSV file.
    """

    # Open participant_details.csv for writing
    with open('participant_details.csv', 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)

        # Write header row to CSV
        writer.writerow(['Handle', 'GeeksForGeeks Handle', 'Codeforces Handle', 'LeetCode Handle', 'CodeChef Handle',
                         'HackerRank Handle', 'GeeksForGeeks URL Exists', 'Codeforces URL Exists',
                         'LeetCode URL Exists',
                         'CodeChef URL Exists', 'HackerRank URL Exists'])

        # Iterate over all files and update participant object details
        for participant in participants:
            with open('geeksforgeeks_handles.txt', 'r') as file:
                for line in file:
                    handle, geeksforgeeks_handle, geeksforgeeks_url_exists = line.split(',')
                    if handle == participant.handle:
                        participant.geeksforgeeks_handle = geeksforgeeks_handle
                        participant.geeksforgeeks_url_exists = geeksforgeeks_url_exists.strip()

            with open('codeforces_handles.txt', 'r') as file:
                for line in file:
                    handle, codeforces_handle, codeforces_url_exists = line.split(',')
                    if handle == participant.handle:
                        participant.codeforces_handle = codeforces_handle
                        participant.codeforces_url_exists = codeforces_url_exists.strip()

            with open('leetcode_handles.txt', 'r') as file:
                for line in file:
                    handle, leetcode_handle, leetcode_url_exists = line.split(',')
                    if handle == participant.handle:
                        participant.leetcode_handle = leetcode_handle
                        participant.leetcode_url_exists = leetcode_url_exists.strip()

            with open('codechef_handles.txt', 'r') as file:
                for line in file:
                    handle, codechef_handle, codechef_url_exists = line.split(',')
                    if handle == participant.handle:
                        participant.codechef_handle = codechef_handle
                        participant.codechef_url_exists = codechef_url_exists.strip()

            with open('hackerrank_handles.txt', 'r') as file:
                for line in file:
                    handle, hackerrank_handle, hackerrank_url_exists = line.split(',')
                    if handle == participant.handle:
                        participant.hackerrank_handle = hackerrank_handle
                        participant.hackerrank_url_exists = hackerrank_url_exists.strip()

            # Write participant details to CSV
            writer.writerow([participant.handle, participant.geeksforgeeks_handle, participant.codeforces_handle,
                             participant.leetcode_handle, participant.codechef_handle, participant.hackerrank_handle,
                             participant.geeksforgeeks_url_exists, participant.codeforces_url_exists,
                             participant.leetcode_url_exists, participant.codechef_url_exists,
                             participant.hackerrank_url_exists])

    # remove all spaces in the csv file
    with open('participant_details.csv', 'r') as file:
        lines = file.readlines()
    with open('participant_details.csv', 'w') as file:
        for line in lines:
            file.write(line.replace(' ', ''))
    # move participant_details.csv to src/main/resources/participant_details.csv, if it exists over write
    os.replace('participant_details.csv', 'src/main/resources/participant_details.csv')        
    print("Participant details written to participant_details.csv")


def main():
    if len(sys.argv) != 3:
        print("Usage: python script.py <excel/csv file path> <platform>")
        return

    file_path = sys.argv[1]
    platform = sys.argv[2].lower()

    platforms = ['geeksforgeeks', 'codeforces', 'leetcode', 'codechef', 'hackerrank']

    if platform not in platforms and platform != 'all' and platform != 'combine':
        print("Invalid platform. Please choose one of: GeeksForGeeks, Codeforces, LeetCode, CodeChef, HackerRank, "
              "All, Combine")
        return

    if not os.path.isfile(file_path):
        print("Invalid file path. Please provide a valid file path.")
        return

    if file_path.endswith('.xlsx'):
        participants = load_excel_sheet(file_path)
    elif file_path.endswith('.csv'):
        participants = load_csv_sheet(file_path)
    else:
        print("Invalid file format. Please provide an Excel (.xlsx) or CSV (.csv) file.")
        return

    if platform == 'geeksforgeeks' or platform == 'all':
        process_geeksforgeeks(participants)
    if platform == 'codeforces' or platform == 'all':
        process_codeforces(participants)
    if platform == 'leetcode' or platform == 'all':
        process_leetcode(participants)
    if platform == 'codechef' or platform == 'all':
        process_codechef(participants)
    if platform == 'hackerrank' or platform == 'all':
        process_hackerrank(participants)
    if platform == 'combine':
        combine_results(participants)


if __name__ == "__main__":
    main()
